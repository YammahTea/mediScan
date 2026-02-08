from fastapi import APIRouter, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse

from io import BytesIO
from typing import List
from datetime import datetime

from copy import copy, deepcopy
import openpyxl

# Modules
from Back.core.pipeline import excel_structure

router = APIRouter(
  tags=["Tools"]
)

@router.post("/tools/merge")
async def merge_files(
        files: List[UploadFile] = File(...)
):
  
  """
  1- Check if the number of files is valid
  2- Loop through all files and checks type and copy headers from first file
  3- Merge the columns
  4- Loop through the new list with the buffer
  5- Save the data
  """
  
  # 1- Check if number of files is valid
  if not files or len(files) < 2:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload at least 2 files to merge")
  
  # main workbook
  merged_wb = openpyxl.Workbook()
  
  # main worksheet
  merged_ws = merged_wb.active
  
  merged_ws.title = "Merged_Patients"
  merged_ws.sheet_view.rightToLeft = True # right to left sheet
  
  current_row_offset = 1 # starts at row 1 (used in the copying data from rows, this is in merged worksheet)
  headers_set = False
  
  try:
    
    for file in files:
      
      # check if type is valid
      if not file.filename.lower().endswith(('.xlsx', '.xls')):
        continue
        
      content = await file.read() # read into memory
      
      # openpyxl
      source_wb = openpyxl.load_workbook(filename=BytesIO(content))
      source_ws = source_wb.active
      
      # 2- copies the headers only from the first file
      if not headers_set:
        for col_index, cell in enumerate(source_ws[1], 1): # first row
          
          new_cell = merged_ws.cell(row=1, column=col_index, value=cell.value)
          
          # extra from me: copy basic styling
          if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            
        # 2.5- adjust the columns' width
        for col_letter, col_dim in source_ws.column_dimensions.items():
          merged_ws.column_dimensions[col_letter].width = col_dim.width
          
        headers_set = True # means that the headers from the first file was copied
      
      
      # 3- copy data rows
      # skip row 1 for all files cuz we already have the column headers
      
      rows = list(source_ws.rows)
      num_rows = len(rows)
      
      if num_rows > 1:
        for row_index, row in enumerate(rows[1:], 1):
          
          target_row_index = current_row_offset + row_index # row index in the merged worksheet
          
          # copies cell values
          for col_index, cell in enumerate(row, 1):
            merged_ws.cell(row=target_row_index, column=col_index, value=cell.value)
            
          # sticker images row height (applies to all)
          merged_ws.row_dimensions[target_row_index].height = source_ws.row_dimensions[row_index + 1].height
          
      # 4- copying images (find images and move them to the new row)
      # note: openpyxl stores images in ws._images (internal) / ws.images
      if hasattr(source_ws, '_images'): # there are images in the source excel file
        for image in source_ws._images:
          
          # Find which row this image belongs to
          # The anchor tells us the top-left cell (e.g., "J5")
          anchor_row = image.anchor._from.row + 1 # the index starts from 0, same thing for the column anchor
          
          # means if an image is in the data range (not header)
          if anchor_row > 1:
            # calc new position
            new_anchor_row = current_row_offset + (anchor_row - 1)
            
            
            # This creates a clone of the image so it doesnt give reference errors or mess up the original reference
            new_img = deepcopy(image)
            
            # anchor of the CLONE
            new_img.anchor._from.row = new_anchor_row - 1
            
            # image height
            height_in_rows = image.anchor.to.row - image.anchor._from.row
            new_img.anchor.to.row = new_anchor_row - 1 + height_in_rows
            
            merged_ws.add_image(new_img)
        
      # update the offset for the next file
      current_row_offset += (num_rows - 1)
    
    # 5- save to buffer
    output = BytesIO()
    merged_wb.save(output)
    output.seek(0)
    
    filename = datetime.now().strftime("Merged_Report_%Y-%m-%d.xlsx")
    
    return StreamingResponse(
      output,
      media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
  
  except Exception as e:
    print(f"Merge error: {e}")
    raise HTTPException(status_code=500, detail="Failed to merge files.")